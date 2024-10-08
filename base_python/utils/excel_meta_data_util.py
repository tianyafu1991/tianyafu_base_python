#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date    : 2022-04-12
# @Author  : tianyafu

from openpyxl.styles import *
import mysqlutil
import gputil
import uuid
import os
import sys

# excel目录的sheet name
catalog_sheet_name = '目录'
# excel中 表信息的sheet模板的sheet name
template_sheet_name = 'Sheet模板'

root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
# 添加自定义class
class_path = root_path + '/meta_data_maintain/classes'
sys.path.append(class_path)

from Model import Model
from TableSheet import TableSheet


def get_tables_from_gp_meta_data(logger, conn):
    """
    通过 sql 从GP中查询出某个db_name 下的所有表以及表的column信息
    :param logger: 日志
    :param db_name: Hive的database name
    :param conn: Hive元数据库的连接
    :return: Hive的db_name库的所有表的元数据信息
    """
    result_dict = dict()
    get_tables_sql = """
SELECT 
       A.TABLENAME AS tbl_name
       ,obj_description(b.oid) tbl_comment
       ,D.ATTNAME AS column_name
       ,REPLACE(REPLACE(REPLACE(FORMAT_TYPE(D.ATTTYPID, D.ATTTYPMOD),'numeric','NUMBER'),'character varying','VARCHAR'),'date','DATE') AS column_type
       ,E.DESCRIPTION column_comment
       ,D.ATTNUM-1 integer_idx
  FROM PG_TABLES A
 INNER JOIN PG_CLASS B
    ON A.TABLENAME = B.RELNAME
  LEFT JOIN PG_CATALOG.PG_DESCRIPTION E
    ON B.OID = E.OBJOID
  LEFT JOIN PG_CATALOG.PG_ATTRIBUTE D
    ON D.ATTRELID = E.OBJOID
   AND D.ATTNUM = E.OBJSUBID
 WHERE SCHEMANAME = 'public'
   AND D.ATTNUM > 0
  ORDER BY A.TABLENAME ,D.ATTNUM
    """
    result = gputil.select_with_conn(logger, get_tables_sql, conn)
    for (tbl_name, tbl_comment, column_name, column_type_name, column_comment, integer_idx) in result:
        tbl_comment = '' if tbl_comment is None else tbl_comment
        column_list = result_dict.get(tbl_name, [])
        table_sheet_instance = TableSheet()
        table_sheet_instance.table_name = tbl_name
        table_sheet_instance.tbl_comment = tbl_comment
        table_sheet_instance.column_name = column_name
        table_sheet_instance.column_type_name = column_type_name
        table_sheet_instance.column_comment = column_comment
        table_sheet_instance.integer_idx = integer_idx
        # column_info = (column_name, column_type_name, column_comment, integer_idx, tbl_comment)
        column_list.append(table_sheet_instance)
        result_dict[tbl_name] = column_list
    logger.info("从元数据库中获取元数据 完成......")
    return result_dict


def get_tables_from_mysql(logger, db_name, conn, need_maintain_table_name_dict):
    result_dict = dict()
    get_tables_sql = """
    SELECT
b.TABLE_NAME AS tbl_name
,a.TABLE_COMMENT as  tbl_comment
,lower(b.COLUMN_NAME) AS column_name
,b.COLUMN_TYPE AS column_type_name
,b.COLUMN_COMMENT AS column_comment
,b.ORDINAL_POSITION AS integer_idx
,0 as table_order
FROM
information_schema.TABLES a
inner join     
information_schema.`COLUMNS` b
on a.table_name = b.table_name
and a.table_schema = b.table_schema
WHERE
b.TABLE_SCHEMA = '%s'
ORDER BY
b.TABLE_NAME,
b.ORDINAL_POSITION
    """ % db_name
    result = mysqlutil.select_with_conn(logger, conn, get_tables_sql)
    for (tbl_name, tbl_comment, column_name, column_type_name, column_comment, integer_idx, table_order) in result:
        if need_maintain_table_name_dict.get(tbl_name, None):
            tbl_comment = '' if tbl_comment is None else tbl_comment
            column_list = result_dict.get(tbl_name, [])
            table_sheet_instance = TableSheet()
            table_sheet_instance.table_name = tbl_name
            table_sheet_instance.tbl_comment = tbl_comment
            table_sheet_instance.column_name = column_name
            table_sheet_instance.column_type_name = column_type_name
            table_sheet_instance.column_comment = column_comment
            table_sheet_instance.integer_idx = integer_idx
            column_list.append(table_sheet_instance)
            result_dict[tbl_name] = column_list
    logger.info("从元数据库中获取元数据 完成......")
    return result_dict


def get_tables_from_hive_meta_data(logger, db_name, conn):
    """
    通过 sql 从Hive的元数据库MySQL中查询出某个db_name下的所有表以及表的column信息
    :param logger: 日志
    :param db_name: Hive的database name
    :param conn: Hive元数据库的连接
    :return: Hive的db_name库的所有表的元数据信息
    """
    result_dict = dict()
    get_tables_sql = """
select
a.tbl_name
,e.param_value tbl_comment
,b.column_name
,b.type_name column_type
,b.`comment` column_comment
,b.integer_idx
,(
case
when a.tbl_name like 'ods%%' then 1
when a.tbl_name like 'dwd%%' then 2
when a.tbl_name like 'dim%%' then 3
when a.tbl_name like 'dws%%' then 4
when a.tbl_name like 'dm%%' then 5
when a.tbl_name like 'app%%' then 6
when a.tbl_name like 'tmp%%' then 7
else 8
end
) table_order
from
tbls a
left join sds d on a.sd_id = d.sd_id 
left join columns_v2 b on d.cd_id = b.cd_id
left join dbs c on a.db_id = c.db_id
left join (select tbl_id,param_value from table_params where param_key = 'comment') e on a.tbl_id = e.tbl_id
where
c.`name` = '%s'
order by
table_order,a.tbl_name,b.integer_idx
    """ % db_name
    result = mysqlutil.select_with_conn(logger, conn, get_tables_sql)
    for (tbl_name, tbl_comment, column_name, column_type_name, column_comment, integer_idx, table_order) in result:
        tbl_comment = '' if tbl_comment is None else tbl_comment
        column_list = result_dict.get(tbl_name, [])
        # column_info = (column_name, column_type_name, column_comment, integer_idx, tbl_comment)
        table_sheet_instance = TableSheet()
        table_sheet_instance.table_name = tbl_name
        table_sheet_instance.tbl_comment = tbl_comment
        table_sheet_instance.column_name = column_name
        table_sheet_instance.column_type_name = column_type_name
        table_sheet_instance.column_comment = column_comment
        table_sheet_instance.integer_idx = integer_idx
        column_list.append(table_sheet_instance)
        result_dict[tbl_name] = column_list
    logger.info("从元数据库中获取元数据 完成......")
    return result_dict


def get_black_list_dict():
    """
    获取黑名单 因部分Hive表仅仅是开发过程中的备份表 或者是 临时表 无需记录元数据 所以需要黑名单过滤
    :return: 不需要维护元数据信息的表的名单 即为黑名单
    """
    black_list_dict = dict()

    black_list_dict['xxx'] = 'xxx'

    return black_list_dict


def parse_catalog_sheet(logger, workbook, sheet_name=catalog_sheet_name):
    """
    解析excel的目录sheet
    :param logger:日志
    :param workbook:excel的workbook实例
    :param sheet_name:需要解析的sheet_name 默认为目录
    :return: 返回excel的目录的内容
    """
    catalog_dict = dict()
    catalog_sheet = workbook[sheet_name]
    if catalog_sheet.max_row == 1:
        return catalog_dict
    else:
        # 因为第一行的是表头 所以min_row 从2开始
        # 用来在for中获取到当前行为哪一行 同样从2开始
        row_index = 2
        for i in catalog_sheet.iter_rows(min_row=2, values_only=True):
            model = Model()
            # 行号
            model.row_no = row_index
            # 编号
            model.serial_number = i[0]
            # 层级
            model.level = i[1]
            if model.level is None:
                # 由于合并单元格导致level为None的 可以通过编号获取
                model.level = str(model.serial_number).split('_')[0]
            # 分类
            model.type = i[2]
            # 表类别
            model.table_type = i[3]
            # 模型名称
            model.table_name = str(i[4]).strip()
            if 'hyperlink' in model.table_name.lower():
                model.table_name = model.table_name.lower().split(',')[1].replace('\"', '').replace(')', '').replace(
                    '\'', '').strip()
            # 模型描述
            model.table_comment = i[5]
            # 调度频率
            model.schedule_frequency = i[6]
            # azkaban调度名称
            model.schedule_name = i[7]
            # 脚本名称
            model.script_name = i[8]
            # 脚本目录
            model.script_dir = i[9]
            # 上线时间
            model.online_time = i[10]
            # 是否采集
            model.collect_flg = i[11]
            # 责任人
            model.principal = i[12]
            # 来源对应表
            model.source_table = i[13]
            # 来源对应表名称
            model.source_table_comment = i[14]
            # 增量/全量
            model.incremental_flg = i[15]
            # 备注
            model.remark = i[16]
            # 超链接对应的sheet 在模型名称(表名)这一列用超链接 链接到了该表具体的sheet中
            hyperlink = catalog_sheet.cell(row_index, 5).hyperlink
            hyperlink_sheet_name = ''
            if hyperlink is not None:
                hyperlink_sheet_name = catalog_sheet.cell(row_index, 5).hyperlink.location
            elif 'hyperlink' in catalog_sheet.cell(row_index, 5).value.lower():
                hyperlink_sheet_name = model.serial_number
            else:
                pass
            # 如果包含! 例如dim_subject_info_yy_i!A1 则去掉!及之后的内容
            hyperlink_sheet_name = hyperlink_sheet_name.split('!')[
                0] if '!' in hyperlink_sheet_name else hyperlink_sheet_name
            # 如果包含'' 则 去掉'及之后的内容
            model.hyperlink_sheet_name = hyperlink_sheet_name[1:len(
                hyperlink_sheet_name) - 1] if '\'' in hyperlink_sheet_name else hyperlink_sheet_name
            catalog_dict[model.table_name] = model
            row_index += 1
    logger.info("解析excel的sheet 完成.......")
    return catalog_dict


def get_table_level(table_name):
    """
    获取表的层级
    :param table_name: 表名
    :return:表的层级
    """
    level = "tmp"
    if table_name.startswith('ods'):
        level = 'ods'
    elif table_name.startswith('dwd'):
        level = 'dwd'
    elif table_name.startswith('dim'):
        level = 'dim'
    elif table_name.startswith('dws'):
        level = 'dws'
    elif table_name.startswith('dm'):
        level = 'dm'
    elif table_name.startswith('app'):
        level = 'app'
    elif table_name.startswith('tmp'):
        level = 'tmp'
    else:
        pass
    return level


def set_border(logger, ws, min_row, max_row, min_col, max_col, style):
    """
    设置边框
    :param logger: 日志
    :param ws: 表的worksheet实例
    :param min_row: 需要边框的最小列
    :param max_row: 需要边框的最大行
    :param min_col: 需要边框的最小列
    :param max_col: 需要边框的最大列
    :param style: 边框的样式
    """
    # 通过 min_row max_row min_col max_col 确定边框区域
    rows = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    # 边框 上下左右四条边
    border_set = Border(left=Side(style=style, color=colors.BLACK),
                        right=Side(style=style, color=colors.BLACK),
                        top=Side(style=style, color=colors.BLACK),
                        bottom=Side(style=style, color=colors.BLACK))
    for row in rows:
        for cell in row:
            cell.border = border_set


def get_gp_tbl_list_dict():
    """
    获取gp中可能需要添加的表 有些表不再Hive中 可能是直接采集到GP中的 这些表需要添加进去
    :return:
    """
    gp_tbl_list_dict = dict()

    gp_tbl_list_dict['xxx'] = 'xxx'

    return gp_tbl_list_dict


def get_missing_meta_data(logger, mysql_dict, catalog_dict, gp_dict, gp_white_dict, hive_black_dict):
    """
    获取在mysql_dict中有 但是excel中缺失的元数据信息
    :param logger: 日志
    :param mysql_dict: Hive元数据库中的元数据信息
    :param catalog_dict: excel目录中的元数据信息
    :param gp_dict: GP数据库中的元数据信息
    :param gp_white_dict: GP数据库中的需要添加的元数据信息 即白名单 因为GP中只有少数表是需要维护元数据到excel中 所以用白名单来标识
    :param hive_black_dict: Hive中不需要添加的元数据信息 即黑名单 因Hive中只有少数表是不需要维护元数据到excel中 所以用黑名单来标识
    :return: 在mysql_dict 但是excel中缺失的元数据信息 和 在gp_dict 但是excel中缺失的元数据信息
    """
    missing_dict = dict()
    # 目录中原来没有数据
    if len(catalog_dict) == 0:
        for table_name in mysql_dict.keys():
            black_info = hive_black_dict.get(table_name, None)
            if black_info is None:
                # 因为目录中原来没有数据 则只要该表不在黑名单中 就是本次需要维护到数据字典中的表
                missing_dict[table_name] = mysql_dict[table_name]
        for tbl_name in gp_dict.keys():
            gp_tbl_info = gp_white_dict.get(tbl_name, None)
            is_add_by_hive = missing_dict.get(tbl_name, None)
            # 是GP中要加入的表 且 在Hive中是不存在的
            if gp_tbl_info is not None and is_add_by_hive is None:
                missing_dict[tbl_name] = gp_dict[tbl_name]
    else:
        for table_name in mysql_dict.keys():
            table_info = catalog_dict.get(table_name, None)
            black_info = hive_black_dict.get(table_name, None)
            # excel目录中没有 并且也不是黑名单
            if table_info is None and black_info is None:
                missing_dict[table_name] = mysql_dict[table_name]
        for tbl_name in gp_dict.keys():
            table_info = catalog_dict.get(tbl_name, None)
            is_add_by_hive = missing_dict.get(tbl_name, None)
            gp_tbl_info = gp_white_dict.get(tbl_name, None)
            # excel目录中没有 且不在Hive中 即只在GP中且需要添加的
            if table_info is None and is_add_by_hive is None and gp_tbl_info is not None:
                missing_dict[tbl_name] = gp_dict[tbl_name]
    logger.info("获取缺失的元数据信息完成.......")
    return missing_dict


def write_catalog_meta_data(logger, level, table_name, table_comment, catalog_sheet, table_no, current_row, table_type):
    """
    写出表的元数据到excel的目录
    :param logger: 日志记录
    :param level: 表层级
    :param table_name: 表名
    :param table_comment: 表注释
    :param catalog_sheet: 目录sheet 的 worksheet实例
    :param table_no: 表的编号 也是表的元数据sheet的sheet name
    :param current_row:
    :param table_type: 表类别
    :return: 该表在目录中的行号
    """
    if current_row is None:
        current_row = catalog_sheet.max_row + 1
    else:
        catalog_sheet.insert_rows(current_row)
    # 编号
    catalog_sheet.cell(current_row, 1).value = table_no
    # 层级
    catalog_sheet.cell(current_row, 2).value = level
    # 模型名称 设置超链接 超链接的格式为 #sheet名!单元格
    # catalog_sheet.cell(current_row, 5).value = table_name
    catalog_sheet.cell(current_row, 5).value = '=HYPERLINK("#\'{}\'!A1", "{}")'.format(table_no, table_name)
    catalog_sheet.cell(current_row, 5).style = "Hyperlink"
    # 模型描述
    catalog_sheet.cell(current_row, 6).value = table_comment
    return current_row


def write_table_sheet_meta_data(logger, table_name, table_comment, table_sheet, table_meta_list, meta_row,
                                border_style):
    """
    写出表的元数据到excel的对应sheet中
    Parameters
    ----------
    logger : 日志
    table_name : 表名
    table_comment : 表注释
    table_sheet : 表的sheet的worksheet实例
    table_meta_list : 表的字段的元数据信息的list
    meta_row : 表在excel目录中的行号
    border_style : excel边框样式

    Returns
    -------

    """
    # 第一行的第二列 即B1 填充表名
    table_sheet.cell(1, 2).value = table_name
    # 第二行的第二列 即B2 填充表注释
    table_sheet.cell(2, 2).value = table_comment
    # 第1行第8列为返回 即H1
    table_sheet.cell(1, 8).value = '=HYPERLINK("#\'{}\'!E{}", "{}")'.format(catalog_sheet_name, str(meta_row), "返回")
    table_sheet.cell(1, 8).style = "Hyperlink"
    # 设置返回这一格的样式
    set_border(logger, table_sheet, 1, 1, 8, 8, border_style)
    # 开始写入字段信息
    for i in range(0, len(table_meta_list)):
        # 字段是从第4行开始写的
        # 字段名
        table_sheet.cell(4 + i, 1).value = table_meta_list[i].column_name
        # 数据类型
        table_sheet.cell(4 + i, 2).value = table_meta_list[i].column_type_name
        # 字段说明
        table_sheet.cell(4 + i, 3).value = table_meta_list[i].column_comment


def write_meta_data_for_init(logger, missing_dict, workbook, style,
                             catalog_sheet_name=catalog_sheet_name, template_sheet_name=template_sheet_name):
    """
    用于初始化excel数据字典时用该方法 因为初始化excel时  只需要将相同level的表写在一起即可 而如果是excel的目录中已有数据 如果要写入新的元数据信息  需要将待插入的表的level与相同的level的数据放在一起
    写出缺失的元数据信息到excel
    :param logger: 日志
    :param missing_dict: 缺失的元数据的字典
    :param workbook: excel的workbook
    :param output_path: 输出的excel文件路径
    :param catalog_sheet_name: excel模板目录sheet的sheet name
    :param template_sheet_name: excel模板中table模板sheet的sheet name
    :param style: 边框样式 默认细边框
    :return:
    """
    catalog_sheet = workbook[catalog_sheet_name]
    template_sheet = workbook[template_sheet_name]
    for table_name in missing_dict.keys():
        table_meta_list = missing_dict[table_name]
        table_comment = table_meta_list[0].tbl_comment
        level = get_table_level(table_name)
        # 生成表的唯一编号
        table_no = level + '_' + str(uuid.uuid4()).split('-')[0]
        table_sheet = workbook.copy_worksheet(template_sheet)
        table_sheet.title = table_no
        # 写目录元数据信息
        meta_row = write_catalog_meta_data(logger, level, table_name, table_comment, catalog_sheet, table_no, None)
        # 写table sheet元数据信息
        write_table_sheet_meta_data(logger, table_name, table_comment, table_sheet, table_meta_list, meta_row, style)
        # min_row=1 第一行 max_row=3 + len(table_meta_list) 字段信息从第4行开始  所以是3+len()  min_col=1 第1列 max_col=7  第G列 为细边框
        set_border(logger, table_sheet, 1, 3 + len(table_meta_list), 1, 7, style)
        logger.info('表%s元数据写出完成' % table_name)


def write_missing_meta_data(catalog_sheet, gp_dict, logger, missing_dict, style, template_sheet, workbook,
                            max_row_no_dict):
    """
    将缺失的数据写入到excel的目录中 并将表字段等元数据信息写入到表的sheet中
    Parameters
    ----------
    catalog_sheet : excel模板目录sheet
    gp_dict : gp表的元数据的字典
    logger : 日志记录
    missing_dict : 要写入的数据的字典
    style : excel边框样式
    template_sheet : excel模板中table模板
    workbook : excel的workbook
    max_row_no_dict : 各个层级的最大的行

    Returns
    -------

    """
    if len(missing_dict) > 0:
        for table_name in missing_dict.keys():
            table_meta_list = missing_dict[table_name]
            # 表注释
            table_comment = table_meta_list[0].tbl_comment
            # 获取表层级
            level = get_table_level(table_name)
            # 生成表的唯一编号
            table_no = level + '_' + str(uuid.uuid4()).split('-')[0]
            # 为该表增加一个sheet
            table_sheet = workbook.copy_worksheet(template_sheet)
            # 该sheet的名称
            table_sheet.title = table_no
            # 表类别
            table_type = "GP" if gp_dict.get(table_name) is not None else "Hive"
            max_row = max_row_no_dict.get(level)
            # 写目录元数据信息
            meta_row = write_catalog_meta_data(logger, level, table_name, table_comment, catalog_sheet, table_no,
                                               max_row,
                                               table_type)
            # 写table sheet元数据信息
            write_table_sheet_meta_data(logger, table_name, table_comment, table_sheet, table_meta_list, meta_row,
                                        style)
            # min_row=1 第一行 max_row=3 + len(table_meta_list) 字段信息从第4行开始  所以是3+len()  min_col=1 第1列 max_col=7  第G列 为细边框
            set_border(logger, table_sheet, 1, 3 + len(table_meta_list), 1, 7, style)
            logger.info('表%s元数据写出完成' % table_name)


def write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, missing_dict, style, template_sheet,
                                            workbook):
    """

    Parameters
    ----------
    catalog_sheet : excel模板目录sheet
    gp_dict : gp表的元数据的字典
    logger : 日志记录
    missing_dict : 要写入的数据的字典
    style : excel边框样式
    template_sheet : excel模板中table模板
    workbook : excel的workbook

    Returns
    -------

    """
    # 获取excel的目录的内容
    excel_catalog_dict = parse_catalog_sheet(logger, workbook)
    # 获取catalog_dict中各个表层级的最大行
    max_row_no_dict = get_level_max_row_no(excel_catalog_dict)
    # 将缺失的数据写入到excel的目录中 并将表字段等元数据信息写入到表的sheet中
    write_missing_meta_data(catalog_sheet, gp_dict, logger, missing_dict, style, template_sheet, workbook,
                            max_row_no_dict)


def write_meta_data_for_maintain(logger, missing_dict, workbook, style, max_row_no_dict, gp_dict,
                                 catalog_sheet_name=catalog_sheet_name, template_sheet_name=template_sheet_name):
    """
    用于维护excel数据字典时用该方法 因为excel的目录中已有数据 如果要写入新的元数据信息  需要将待插入的表的level与相同的level的数据放在一起
    写出缺失的元数据信息到excel
    Parameters
    ----------
    logger : 日志
    missing_dict : 缺失的元数据的字典
    workbook : excel的workbook
    style : 边框样式 默认细边框
    max_row_no_dict : 各个层级的最大的行
    gp_dict : gp表的元数据的字典
    catalog_sheet_name : excel模板目录sheet的sheet name
    template_sheet_name : excel模板中table模板sheet的sheet name

    Returns
    -------

    """
    catalog_sheet = workbook[catalog_sheet_name]
    template_sheet = workbook[template_sheet_name]
    ods_missing_dict = dict()
    dwd_missing_dict = dict()
    dim_missing_dict = dict()
    dws_missing_dict = dict()
    dm_missing_dict = dict()
    app_missing_dict = dict()
    tmp_missing_dict = dict()
    for tbl_name in missing_dict.keys():
        level = get_table_level(tbl_name)
        # level = "tmp"
        if level == 'ods':
            ods_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'dwd':
            dwd_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'dim':
            dim_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'dws':
            dws_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'dm':
            dm_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'tmp':
            tmp_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        elif level == 'app':
            app_missing_dict[tbl_name] = missing_dict.get(tbl_name)
        else:
            tmp_missing_dict[tbl_name] = missing_dict.get(tbl_name)
    # 写入缺失的ods表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, ods_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的dwd表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, dwd_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的dim表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, dim_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的dws表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, dws_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的dm表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, dm_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的app表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, app_missing_dict, style, template_sheet,
                                            workbook)
    # 写入缺失的tmp表
    write_missing_meta_data_by_missing_dict(catalog_sheet, gp_dict, logger, tmp_missing_dict, style, template_sheet,
                                            workbook)


def get_level_max_row_no(catalog_dict):
    """
    获取catalog_dict中各个表层级的最大行
    :param catalog_dict:目录的元数据dict
    :return: 各个层级的最大行号的dict
    """
    result_dict = {}
    for model in catalog_dict.values():
        row_no = model.row_no
        level = model.level
        dict_row_no = result_dict.get(level)
        if dict_row_no is None:
            # result_dict中没有该层级 则将本model的行号放入
            result_dict[level] = row_no
        else:
            # 本model的行号比result_dict中的该level的行号大时 取本model的行号
            result_dict[level] = row_no if row_no > dict_row_no else dict_row_no
    return result_dict


def maintain_meta_data(logger, workbook, mysql_dict, catalog_dict, style, gp_dict):
    catalog_sheet = workbook[catalog_sheet_name]
    for model in catalog_dict.values():
        catalog_row_no = model.row_no
        table_name = model.table_name
        table_sheet = workbook[model.serial_number]
        # 如果是Hive表  或者 Hive&GP 或者 Hive&MySQL  统一取Hive表的元数据
        table_meta_list = mysql_dict.get(table_name, None)
        # 有些表是单纯的GP表 这时候应该从GP的元数据中查
        if 'gp' == str(model.table_type).lower():
            table_meta_list = gp_dict.get(table_name, None)
        if table_meta_list is None:
            continue
        else:
            # 这里直接暴力点 直接将excel中原有的字段信息删除掉后 再写入 不采取有变化才更新数据
            # 先要删除原来的内容 这里采取倒着删  因为正着删时  删除行后下面的行会往上移所以容易出现没删干净的情况  如果要正着删  就是一直删除第4行
            for i in range(table_sheet.max_row, 3, -1):
                table_sheet.delete_rows(i)
            # 写入新的元数据信息
            # 表的注释可能发生变化 所以连注释一起更新掉
            table_comment = ''
            for i in range(0, len(table_meta_list)):
                # 字段信息从第4行开始写
                table_sheet.cell(4 + i, 1).value = table_meta_list[i].column_name
                table_sheet.cell(4 + i, 2).value = table_meta_list[i].column_type_name
                table_sheet.cell(4 + i, 3).value = table_meta_list[i].column_comment
                if i == 0:
                    table_comment = table_meta_list[i].tbl_comment
            # 维护表的注释
            table_sheet.cell(2, 2).value = table_comment
            # 维护表的sheet的返回的超链接  因为可能目录发生了变化  导致原先返回的超链接不准确了
            table_sheet.cell(1, 8).value = '=HYPERLINK("#\'{}\'!E{}", "{}")'.format(catalog_sheet_name,
                                                                                    str(catalog_row_no),
                                                                                    "返回")
            table_sheet.cell(1, 8).style = "Hyperlink"
            # 同步更新掉目录中的表的注释
            catalog_sheet.cell(catalog_row_no, 6).value = table_comment

            # 维护sheet边框
            set_border(logger, table_sheet, 1, 3 + len(table_meta_list), 1, 7, style)


def unmerge_catalog_cells(logger, workbook):
    """
    将层级的合并单元格拆分掉
    Parameters
    ----------
    logger
    workbook : 文档对象

    Returns
    -------

    """
    # 获取目录的sheet
    catalog_sheet = workbook[catalog_sheet_name]
    # merged_ranges = catalog_sheet.merged_cells.ranges
    # 这个merged_ranges是一个list
    merged_ranges = catalog_sheet.merged_cell_ranges
    for merged_range in merged_ranges:
        # merged_range_str例如B2:B31
        merged_range_str = str(merged_range)
        logger.info("有合并单元格%s" % merged_range)
        # 只有当合并单元格是由2个及以上的单元格合并而来 才需要拆分
        if len(merged_range_str.split(':')) == 2 and merged_range_str.split(':')[0].startswith("B") and \
                merged_range_str.split(':')[1].startswith("B"):
            catalog_sheet.unmerge_cells(merged_range_str)


def merge_catalog_cells(logger, catalog_dict, catalog_sheet):
    min_max_dict = dict()
    for model in catalog_dict.values():
        level = str(model.serial_number).split('_')[0]
        level_list = min_max_dict.get(level, [])
        level_list.append(model.row_no)
        min_max_dict[level] = level_list
    for level in min_max_dict.keys():
        level_list = min_max_dict.get(level)
        max_row = max(level_list)
        min_row = min(level_list)
        print(min_row, max_row, type(min_row), type(max_row))
        # logger.info("min:" + str(min_row) + "max:" +  str(max_row) + "!!!" + type(min_row) + "$$$$$$$$" + type(max_row))
        catalog_sheet.merge_cells(start_row=min_row, start_column=2, end_row=max_row, end_column=2)
        logger.info("合并目录层级为%s的 从%s行合并到%s行" % (level, min_row, max_row))
    logger.info("目录的层级合并成功......")


def catalog_merge_level(logger, catalog_sheet):
    """
    合并层级
    :param logger: 日志
    :param catalog_sheet: 目录的worksheet实例
    :return:
    """
    level_dict = dict()
    column = 2
    # 第2列为层级 相同的层级的单元格可以合并 所以这里col都是2  min_row为2 因第一行是表头
    for row in catalog_sheet.iter_rows(min_row=2, max_row=catalog_sheet.max_row, min_col=1, max_col=1):
        level = str(row[0].value).split('_')[0]
        row_num = row[0].row
        level_list = level_dict.get(level, [])
        level_list.append(row_num)
        level_dict[level] = level_list
    for level in level_dict.keys():
        level_list = level_dict[level]
        max_row = max(level_list)
        min_row = min(level_list)
        # catalog_sheet.merge_cells('{}{}:{}{}'.format(column, str(min_row), column, str(max_row)))

        top_left_cell_point = 'B%s' % min_row
        top_left_cell = catalog_sheet[top_left_cell_point]
        top_left_cell.value = level
        catalog_sheet.merge_cells(start_row=min_row, start_column=column, end_row=max_row, end_column=column)
        logger.info("合并目录层级为%s的 从%s行合并到%s行" % (level, min_row, max_row))
    logger.info("目录的层级合并成功......")
