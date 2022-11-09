#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date    : 2022-04-12
# @Author  : tianyafu

"""
用于初始化数据字典 只初始化Hive表的元数据信息  因为MySQL表中存在很多是后端的业务表  而GP表大部分都是Hive下发过去的(除了算法可能是直接写GP的 以及 可能会有部分其他表是只在GP的)
项目前期开发时 不需要花很多精力用来维护数据字典 在开发之后 一次性初始化数据字典
后续表结构等元数据信息可能会发生变化 这个不是本脚本的功能 用
"""

import os
import sys
import configparser
import datetime
import openpyxl

# 定义root path
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))

# 添加自定义模块到系统路径
mylib_path = root_path + '/utils'
sys.path.append(mylib_path)
import mysqlutil
from logutil import Logging
import excel_meta_data_util
import gputil

# 获取日志logger
logger = Logging().get_logger()

# 加载配置文件
config_path = root_path + '/config/prod_dev.conf'
config = configparser.ConfigParser()
config.read(config_path, encoding="utf-8-sig")
# MySQL连接信息
hive_mysql_section = "hive_mysql"
hive_mysql_host = config.get(hive_mysql_section, "host")
hive_mysql_user = config.get(hive_mysql_section, "user")
hive_mysql_passwd = config.get(hive_mysql_section, "passwd")
hive_mysql_database = config.get(hive_mysql_section, "database")
hive_mysql_port = config.get(hive_mysql_section, "port")
# 项目的信息
project_section = "project"
project_name = config.get(project_section, "name")
project_hive_db = config.get(project_section, "hive_db")
# excel目录的sheet name
catalog_sheet_name = config.get(project_section, "catalog_sheet")
# excel中 表信息的sheet模板的sheet name
template_sheet_name = config.get(project_section, "template_sheet")
# excel边框样式
border_style = config.get(project_section, "border_style")

gp_section = "greenplum"
gp_host = config.get(gp_section, "host")
gp_user = config.get(gp_section, "user")
gp_passwd = config.get(gp_section, "passwd")
gp_database = config.get(gp_section, "database")
gp_port = config.get(gp_section, "port")


def catalog_merge_level(logger, catalog_sheet):
    """
    合并层级
    :param logger: 日志
    :param catalog_sheet: 目录的worksheet实例
    :return:
    """
    level_dict = dict()
    column = 2
    # 第2列为层级 相同的层级的单元格可以合并 所以这里col都是2  min_row为2 因第一行是表头
    for row in catalog_sheet.iter_rows(min_row=2, max_row=catalog_sheet.max_row, min_col=2, max_col=2):
        level = row[0].value
        row_num = row[0].row
        column = row[0].column
        level_list = level_dict.get(level, [])
        level_list.append(row_num)
        level_dict[level] = level_list
    for level in level_dict.keys():
        level_list = level_dict[level]
        max_row = max(level_list)
        min_row = min(level_list)
        # catalog_sheet.merge_cells('{}{}:{}{}'.format(column, str(min_row), column, str(max_row)))
        catalog_sheet.merge_cells(start_row=min_row, start_column=column, end_row=max_row, end_column=column)
    logger.info("目录的层级合并成功......")


if __name__ == '__main__':
    input_path = r'../input/数据模型-模板.xlsx'
    output_file_name = "%s数据字典@%s_init.xlsx" % (project_name, datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
    output_path = "../output/%s" % output_file_name
    # excel模板的绝对路径
    abspath = os.path.abspath(input_path)
    try:
        # excel模板的workbook实例
        workbook = openpyxl.load_workbook(abspath)
        # 连接Hive元数据库
        mysql_conn = mysqlutil.connect_with_port(logger, hive_mysql_host, hive_mysql_user, hive_mysql_passwd,
                                                 hive_mysql_database,
                                                 int(hive_mysql_port))
        # gp连接
        gp_conn = gputil.maintain_conn(logger, None, gp_host, gp_user, gp_passwd, gp_database, gp_port)
        # 从Hive元数据库中 查询出表的元数据
        mysql_result_dict = excel_meta_data_util.get_tables_from_hive_meta_data(logger, project_hive_db, mysql_conn)
        gp_dict = excel_meta_data_util.get_tables_from_gp_meta_data(logger, gp_conn)
        # 解析excel模板 将目录中的所有内容解析出来
        catalog_dict = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        # 获取excel模板中缺失的元数据信息
        missing_dict = excel_meta_data_util.get_missing_meta_data(logger, mysql_result_dict, catalog_dict, gp_dict)
        # 写出缺失的元数据信息到excel
        excel_meta_data_util.write_meta_data_for_init(logger, missing_dict, workbook, border_style)
        catalog_sheet = workbook[catalog_sheet_name]
        # 第2列为层级 相同的层级的单元格可以合并
        catalog_merge_level(logger, catalog_sheet)
        # 目录添加边框
        excel_meta_data_util.set_border(logger, catalog_sheet, catalog_sheet.min_row, catalog_sheet.max_row,
                                        catalog_sheet.min_column,
                                        catalog_sheet.max_column, border_style)
        # 写出到文件
        workbook.save(output_path)
    except Exception as e:
        logger.error(e)
        raise Exception
    finally:
        mysqlutil.close(mysql_conn)
