#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date    : 2022-04-11
# @Author  : tianyafu


"""
openpyxl 库的文档见:https://openpyxl.readthedocs.io/en/stable/
openpyxl 库的安装: pip3 install openpyxl==3.0.10

因为部分项目只需要用到MySQL 所以只需要维护MySQL的元数据
"""

import os
import sys
import configparser
import datetime

# 添加自定义模块到系统路径

mylib_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)) + '/utils'
sys.path.append(mylib_path)
import mysqlutil
from logutil import Logging
import openpyxl
import excel_meta_data_util

# 获取日志logger
logger = Logging().get_logger()

# 加载配置文件
config_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)) + '/config/prod_dev.conf'
config = configparser.ConfigParser()
config.read(config_path, encoding="utf-8-sig")
# 整个项目以MySQL为主
mysql_section = "front_mysql"
mysql_host = config.get(mysql_section, "host")
mysql_user = config.get(mysql_section, "user")
mysql_passwd = config.get(mysql_section, "passwd")
mysql_database = config.get(mysql_section, "database")
mysql_port = config.get(mysql_section, "port")

# 项目的信息
project_section = "project"
project_name = config.get(project_section, "name")
# excel目录的sheet name
catalog_sheet_name = config.get(project_section, "catalog_sheet")
# excel中 表信息的sheet模板的sheet name
template_sheet_name = config.get(project_section, "template_sheet")
# excel边框样式
border_style = config.get(project_section, "border_style")


def get_mysql_table_name_dict():
    """
    获取要维护元数据的表名
    :return:
    """
    result_dict = {}
    table_list = []

    # table_list.append("xxx")

    for i in table_list:
        result_dict[i] = i
    return result_dict


def get_gp_white_tbl_list_dict():
    """
    获取gp中可能需要添加的表 有些表不再Hive中 可能是直接采集到GP中的 这些表需要添加进去
    :return:
    """
    gp_tbl_list_dict = dict()

    # gp_tbl_list_dict['xxx'] = 'xxx'

    return gp_tbl_list_dict


def get_hive_black_list_dict():
    """
    获取黑名单 因部分Hive表仅仅是开发过程中的备份表 或者是 临时表 无需记录元数据 所以需要黑名单过滤
    :return: 不需要维护元数据信息的表的名单 即为黑名单
    """
    black_list_dict = dict()

    # black_list_dict['xxx'] = 'xxx'

    return black_list_dict


if __name__ == '__main__':
    try:
        excel_input_path = r'../input/数据模型-模板.xlsx'
        output_file_name = "%s数据字典@%s.xlsx" % (project_name, datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        output_path = "../output/%s" % output_file_name
        abspath = os.path.abspath(excel_input_path)
        workbook = openpyxl.load_workbook(abspath)
        # 先将层级的合并单元格拆分掉
        excel_meta_data_util.unmerge_catalog_cells(logger, workbook)
        mysql_conn = mysqlutil.connect_with_port(logger, mysql_host, mysql_user, mysql_passwd,
                                                 mysql_database,
                                                 int(mysql_port))
        # 获取要维护元数据的表名
        need_maintain_table_name_dict = get_mysql_table_name_dict()
        # 通过 sql 从MySQL中查询出某个db_name下的所有表以及表的column信息
        mysql_dict = excel_meta_data_util.get_tables_from_mysql(logger, mysql_database, mysql_conn,
                                                                need_maintain_table_name_dict)
        gp_dict = dict()
        # 获取excel的目录的内容
        excel_catalog_dict = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        # 获取excel中缺失的元数据信息 因为只需要维护MySQL的元数据 所以后3个参数都给了个空dict
        missing_dict = excel_meta_data_util.get_missing_meta_data(logger, mysql_dict, excel_catalog_dict, gp_dict,
                                                                  gp_dict, gp_dict)
        # 有excel目录中缺失的元数据信息 才进行缺失元数据的写入
        if len(missing_dict) > 0:
            # 获取各个层级的最大的行
            level_max_row_no_dict = excel_meta_data_util.get_level_max_row_no(excel_catalog_dict)
            # 将缺失的元数据信息写入到workbook 包括写目录和写对应的sheet
            excel_meta_data_util.write_meta_data_for_maintain(logger, missing_dict, workbook, border_style,
                                                              level_max_row_no_dict, gp_dict)
        # 因为上面的if中 可能会对excel的目录发生改变 导致原先获取的excel_catalog_dict中的行号是不准的 所以需要重新获取excel_catalog_dict
        excel_catalog_dict_2 = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        # 维护所有sheet的字段信息元数据
        excel_meta_data_util.maintain_meta_data(logger, workbook, mysql_dict, excel_catalog_dict_2, border_style,
                                                gp_dict)
        # 需要重新维护目录的合并单元格和边框
        # 目录添加边框
        catalog_sheet = workbook[catalog_sheet_name]
        excel_meta_data_util.set_border(logger, catalog_sheet, catalog_sheet.min_row, catalog_sheet.max_row,
                                        catalog_sheet.min_column,
                                        catalog_sheet.max_column, border_style)
        # 重新获取excel的目录的内容 因为目录的信息已经过重新维护 需要重新进行单元格合并
        maintained_excel_catalog_dict = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        excel_meta_data_util.catalog_merge_level(logger, catalog_sheet)
        workbook.save(output_path)
    except Exception as e:
        logger.error(e)
        raise Exception
    finally:
        mysqlutil.close(mysql_conn)
