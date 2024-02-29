#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date    : 2022-04-11
# @Author  : tianyafu


"""
openpyxl 库的文档见:https://openpyxl.readthedocs.io/en/stable/
"""

import os
import sys
import configparser
import datetime

# 定义root path
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))

# 添加自定义模块到系统路径
mylib_path = root_path + '/utils'
sys.path.append(mylib_path)
import mysqlutil
import gputil
from logutil import Logging
import openpyxl
import excel_meta_data_util

# 获取日志logger
logger = Logging().get_logger()

# 加载配置文件
config_path = root_path + '/config/prod_dev.conf'
config = configparser.ConfigParser()
config.read(config_path, encoding="utf-8-sig")
hive_mysql_section = "hive_mysql"
hive_mysql_host = config.get(hive_mysql_section, "host")
hive_mysql_user = config.get(hive_mysql_section, "user")
hive_mysql_passwd = config.get(hive_mysql_section, "passwd")
hive_mysql_database = config.get(hive_mysql_section, "database")
hive_mysql_port = config.get(hive_mysql_section, "port")

# 项目的信息
project_section = "project"
project_name = config.get(project_section, "name")
project_hive_db = config.get(project_section, "hive_db")
# excel目录的sheet name
catalog_sheet_name = config.get(project_section, "catalog_sheet")
# excel中 表信息的sheet模板的sheet name
template_sheet_name = config.get(project_section, "template_sheet")
# excel边框样式
border_style = config.get(project_section, "border_style")

gp_section = "greenplum"
gp_host = config.get(gp_section, "host")
gp_user = config.get(gp_section, "user")
gp_passwd = config.get(gp_section, "passwd")
gp_database = config.get(gp_section, "database")
gp_port = config.get(gp_section, "port")


def get_sheet_content(main_book):
    """
    获取表名的超链接的sheet_name
    因部分表名长度较长  用表名做sheet_name时 会提示超出长度  所以表名与超链接的sheet_name 并不是一一对应的 所以这里要先获取hyperlink_sheet_name
    """
    table_sheet_dict = dict()
    catalog_sheet = main_book['目录']
    # 因range函数为左闭右开 所以这里catalog_sheet.max_row 需要 加 1  而第一行为表头 所以从2开始
    for i in range(2, catalog_sheet.max_row + 1):
        table_name = catalog_sheet.cell(i, 5).value
        hyperlink_sheet_name = catalog_sheet.cell(i, 5).hyperlink.location
        if '!' in hyperlink_sheet_name:
            hyperlink_sheet_name = hyperlink_sheet_name.split('!')[0]
            if '\'' in hyperlink_sheet_name:
                hyperlink_sheet_name = hyperlink_sheet_name[1:len(hyperlink_sheet_name) - 1]
        table_sheet_content = main_book[hyperlink_sheet_name]
        column_list = []
        max_row = table_sheet_content.max_row
        # 有可能这个sheet中没有内容 只有表头 这时候需要在这里添加一行 用来将该table的超链接信息传递出去
        if max_row < 4:
            column_list.append(('', '', '', hyperlink_sheet_name))
            table_sheet_dict[table_name] = column_list
        else:
            for i in range(4, table_sheet_content.max_row + 1):
                column_name = table_sheet_content.cell(i, 1).value
                column_type_name = table_sheet_content.cell(i, 2).value
                column_comment = table_sheet_content.cell(i, 3).value
                column_info = (column_name, column_type_name, column_comment, hyperlink_sheet_name)
                column_list.append(column_info)
            table_sheet_dict[table_name] = column_list
    return table_sheet_dict

    # main_book.save("1.xlsx")


if __name__ == '__main__':
    try:
        excel_input_path = r'../input/xxxx.xlsx'
        output_file_name = "%s数据字典@%s.xlsx" % (project_name, datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        output_path = "../output/%s" % output_file_name
        abspath = os.path.abspath(excel_input_path)
        workbook = openpyxl.load_workbook(abspath)
        hive_mysql_conn = mysqlutil.connect_with_port(logger, hive_mysql_host, hive_mysql_user, hive_mysql_passwd,
                                                      hive_mysql_database,
                                                      int(hive_mysql_port))
        gp_conn = gputil.maintain_conn(logger, None, gp_host, gp_user, gp_passwd, gp_database, gp_port)
        # 先将层级的合并单元格拆分掉
        excel_meta_data_util.unmerge_catalog_cells(logger, workbook)
        # 通过 sql 从Hive的元数据库MySQL中查询出某个db_name下的所有表以及表的column信息
        mysql_dict = excel_meta_data_util.get_tables_from_hive_meta_data(logger, project_hive_db, hive_mysql_conn)
        # 通过 sql 从GP中查询出某个db_name 下的所有表以及表的column信息
        gp_dict = excel_meta_data_util.get_tables_from_gp_meta_data(logger, gp_conn)
        # 获取excel的目录的内容
        excel_catalog_dict = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        # 获取excel中缺失的元数据信息
        missing_dict = excel_meta_data_util.get_missing_meta_data(logger, mysql_dict, excel_catalog_dict, gp_dict)
        # 有excel目录中缺失的元数据信息 才进行缺失元数据的写入
        if len(missing_dict) > 0:
            # 获取各个层级的最大的行
            level_max_row_no_dict = excel_meta_data_util.get_level_max_row_no(excel_catalog_dict)
            # 将缺失的元数据信息写入到workbook 包括写目录和写对应的sheet
            excel_meta_data_util.write_meta_data_for_maintain(logger, missing_dict, workbook, border_style,
                                                              level_max_row_no_dict, gp_dict)
        # 因为上面的if中 可能会对excel的目录发生改变 导致原先获取的excel_catalog_dict中的行号是不准的 所以需要重新获取excel_catalog_dict
        excel_catalog_dict_2 = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        # 维护所有sheet的字段信息元数据
        excel_meta_data_util.maintain_meta_data(logger, workbook, mysql_dict, excel_catalog_dict_2, border_style,
                                                gp_dict)
        # 需要重新维护目录的合并单元格和边框
        # 目录添加边框
        catalog_sheet = workbook[catalog_sheet_name]
        excel_meta_data_util.set_border(logger, catalog_sheet, catalog_sheet.min_row, catalog_sheet.max_row,
                                        catalog_sheet.min_column,
                                        catalog_sheet.max_column, border_style)
        # 重新获取excel的目录的内容 因为目录的信息已经过重新维护 需要重新进行单元格合并
        maintained_excel_catalog_dict = excel_meta_data_util.parse_catalog_sheet(logger, workbook)
        excel_meta_data_util.catalog_merge_level(logger, catalog_sheet)
        workbook.save(output_path)
    except Exception as e:
        logger.error(e)
        raise Exception
    finally:
        mysqlutil.close(hive_mysql_conn)
        gputil.close(gp_conn)
