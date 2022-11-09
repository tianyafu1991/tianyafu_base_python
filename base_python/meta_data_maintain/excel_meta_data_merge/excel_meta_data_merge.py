#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Date    : 2022-04-11
# @Author  : tianyafu


"""
用于将一份已有的数据字典的目录中的其他信息 拷贝到另一份excel数据字典的目录中
因省信访原有一份手工维护的数据字典 后续用excel_meta_data_init脚本生成了一份excel数据字典
但是脚本生成的数据字典中  只有Hive元数据库中有的信息  缺少了调度 负责人  表类型等其他需要手工维护的信息
所以需要把 手工维护的excel中已维护好的信息  按照表名 拷贝到脚本生成的excel数据字典中
"""

import os
import sys
import datetime
import configparser


# 定义root path
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))

# 添加自定义模块到系统路径
mylib_path = root_path + '/utils'
sys.path.append(mylib_path)
import excel_meta_data_util
from logutil import Logging
import openpyxl

# 获取日志logger
logger = Logging().get_logger()

# 加载配置文件
config_path = root_path + '/config/prod_dev.conf'
config = configparser.ConfigParser()
config.read(config_path, encoding="utf-8-sig")

project_section = "project"
project_name = config.get(project_section, "name")


def get_workbook(logger, path):
    """
    获取path对应的excel的workbook实例
    :param logger 日志
    :param path excel的相对路径
    :return: excel的workbook实例
    """
    abspath = os.path.abspath(path)
    workbook = openpyxl.load_workbook(abspath)
    logger.info("获取excel:%s的workbook实例成功......" % abspath)
    return workbook


def merge_meta_data(logger, manual_dict, workbook):
    """
    将手工维护好的excel中的其他元数据信息更新到脚本生成的excel的workbook实例 因手工维护好的excel中包含了调度  责任人 增量全量等Hive元数据中无法获取到的信息 所以需要有这一步
    :param logger: 日志
    :param manual_dict:手工维护好的excel的dict
    :param workbook:脚本生成的excel的workbook实例
    """
    catalog_sheet = workbook[excel_meta_data_util.catalog_sheet_name]
    for row in catalog_sheet.iter_rows(min_row=2, max_row=catalog_sheet.max_row):
        table_name = str(row[4].value).strip()
        if 'hyperlink' in table_name.lower():
            table_name = table_name.lower().split(',')[1].replace('\"', '').replace(')', '').replace('\'', '').strip()
        manual_model = manual_dict.get(table_name, None)
        if manual_model is None:
            logger.warning("表%s在手工维护的excel中没有找到" % table_name)
            continue
        # 分类
        row[2].value = manual_model.type
        # 表类别
        row[3].value = manual_model.table_type
        # 调度频率
        row[6].value = manual_model.schedule_frequency
        # azkaban调度名称
        row[7].value = manual_model.schedule_name
        # 脚本名称
        row[8].value = manual_model.script_name
        # 脚本目录
        row[9].value = manual_model.script_dir
        # 上线时间
        row[10].value = manual_model.online_time
        # 是否采集
        row[11].value = manual_model.collect_flg
        # 责任人
        row[12].value = manual_model.principal
        # 来源对应表
        row[13].value = manual_model.source_table
        # 来源对应表名称
        row[14].value = manual_model.source_table_comment
        # 增量/全量
        row[15].value = manual_model.incremental_flg
        # 备注
        row[16].value = manual_model.remark
    logger.info("合并信息成功......")


if __name__ == '__main__':
    try:
        output_file_name = "%s数据字典@%s_merge.xlsx" % (project_name, datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
        output_path = "../output/%s" % output_file_name
        # 手工维护的excel
        manual_maintenance_input_path = r'../input/xxx.xlsx'
        manual_maintenance_workbook = get_workbook(logger, manual_maintenance_input_path)
        manual_dict = excel_meta_data_util.parse_catalog_sheet(logger, manual_maintenance_workbook)
        # 脚本维护的excel
        script_maintenance_input_path = r'../input/xxxx.xlsx'
        script_maintenance_workbook = get_workbook(logger, script_maintenance_input_path)
        # 合并2份excel的元数据
        merge_meta_data(logger, manual_dict, script_maintenance_workbook)
        # 保存成文件
        script_maintenance_workbook.save(output_path)
        logger.info("处理完成 保存文件成功......")
    except Exception as e:
        logger.error(e)
        raise Exception
    finally:
        pass
